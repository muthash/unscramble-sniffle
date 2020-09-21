import json
import os
from openpyxl import load_workbook
# from bop_json_convert2 import insert_cell_value


file_dir = os.path.dirname(os.path.abspath(__file__))
# wb = load_workbook(os.path.join(file_dir, "excels/BOP.xlsx"), read_only=True)
# wb2 = load_workbook(os.path.join(file_dir,"excels/WC.xlsx"), read_only=True)
wb3 = load_workbook(os.path.join(file_dir, "excels/GL.xlsx"), read_only=True)

if not wb:
    print('Not supposed to get here')
    exit()


def bop_data():
    return {
        "description": "Multiclass - Wholesaler OH",
        "via": "CHROME",
        "featureSets": [
            {
                "description": "Create ACCOUNT - Wholesaler OH",
                "testCasetype": "ACCOUNT",
                "execute": True,
                "testData": {
                    "address": {
                        "addressLine1": "150 W. Main Street",
                        "city": "New Albany",
                        "state": "OH",
                        "zip": "43054"
                    },
                    "name": "Wholesaler OH",
                    "naicsCodes": [
                        {
                            "code": "423720",
                            "description": "Gas hot water heaters merchant wholesalers"
                        }
                    ],
                    "insuredPhone": "6092294392",
                    "insuredEmail": "davenmathews@gmail.com",
                    "organizationType": "corporation",
                    "yearsInBusiness": 4
                }
            },
            {
                "description": "Quote BOP - Wholesaler OH",
                "testCasetype": "QUOTE",
                "execute": True,
                "expect": {
                    "status": "quoted"
                },
                "testData": {
                    "insuranceProduct": "BOP",
                    "locations": [
                        {
                            "employeeCount": 24,
                            "propertyDeductible": 1000,
                            "buildings": [
                                {
                                    "exposure": {
                                        "lessorsRisk": False,
                                        "businessType": "Wholesaler",
                                        "classification": {
                                            "code": "50581",
                                            "descriptionCode": "AC /Combined AC & Heating Equipment-Distbtrs Only"
                                        },
                                        "limitForBusinessPersonalProperty": 25000,
                                        "constructionType": "JoistedMasonry",
                                        "storyCount": "3",
                                        "squareFootage": 10000,
                                        "yearBuilt": 1988,
                                        "hasAutomaticSprinklerSystem": False,
                                        "buildingLimit": 100000,
                                        "roofUpdated": True,
                                        "electricPlumbingHVACUpdated": True
                                    },
                                    "coverage": {
                                        "utilityServicesDirectDamage": 0,
                                        "utilityServicesTimeElement": 0,
                                        "businessIncomeAndExtraExpensesIndemnityInMonths": 9,
                                        "windCoverage": {
                                            "windDeductiblePercent": "10%"
                                        },
                                        "equipmentBreakdownCoverage": False,
                                        "ordinanceAndLawCoverage": 10000,
                                        "debrisRemoval": 25000
                                    }
                                }
                            ]
                        }
                    ],
                    "policy": {
                        "coverages": [
                            {
                                "name": "PerOccurrence",
                                "value": {
                                    "occurrenceLimit": 500000
                                }
                            },
                            {
                                "name": "PerPersonMedical",
                                "value": {
                                    "perPersonLimit": 5000
                                }
                            },
                            {
                                "name": "EmployeeBenefits",
                                "value": {
                                    "aggregateLimit": 500000,
                                    "deductible": 0,
                                    "perEmployeeLimit": 500000,
                                    "retroactiveDate": {
                                        "numberOfDaysInFuture": 0
                                    }
                                }
                            },
                            {
                                "name": "DamageToPremises",
                                "value": {
                                    "occurrenceLimit": 400000
                                }
                            },
                            {
                                "name": "HiredNonOwnedAuto",
                                "value": {
                                    "optedIn": True
                                }
                            },
                            {
                                "name": "EmployeeDishonesty",
                                "value": {
                                    "aggregateLimit": 25000
                                }
                            },
                            {
                                "name": "EmployeeRelatedPracticesLiability",
                                "value": {
                                    "aggregateLimit": 10000,
                                    "deductible": 5000,
                                    "perEmploymentWrongfulActLimit": 10000,
                                    "retroactiveDate": {
                                        "numberOfDaysInFuture": 0
                                    }
                                }
                            },
                            {
                                "name": "WaiverOfSubrogation",
                                "value": {
                                    "waiversOfSubrogation": [
                                        {
                                            "orgOrPersonName": "Dave Mathews"
                                        },
                                        {
                                            "orgOrPersonName": "Hemal Patel"
                                        }
                                    ]
                                }
                            }
                        ],
                        "effectiveDate": {
                            "numberOfDaysInFuture": 0
                        }
                    }
                }
            }
        ]
    }


def gl_data():
    return {
        "description": "Multiclass - Category State",
        "via": "CHROME",
        "featureSets": [
            {
                "description": "Create ACCOUNT - Account Name",
                "testCasetype": "ACCOUNT",
                "execute": True,
                "testData": {
                    "address": {
                        "addressLine1": "150 W. Main Street",
                        "city": "New Albany",
                        "state": "OH",
                        "zip": "43054"
                    },
                    "name": "Wholesaler OH",
                    "naicsCodes": [
                        {
                            "code": "423720",
                            "description": "Gas hot water heaters merchant wholesalers"
                        }
                    ],
                    "insuredPhone": "6092294392",
                    "insuredEmail": "davenmathews@gmail.com",
                    "organizationType": "corporation",
                    "yearsInBusiness": 4
                }
            },
            {
                "description": "Quote GL - Category State",
                "testCasetype": "QUOTE",
                "execute": True,
                "expect": {
                    "status": "quoted"
                },
                "testData": {
                    "requestUuid": "HISCOX_REQUEST_UUID",
                    "accountId": "0243121244",
                    "quoteId": "",
                    "classOfBusinessCd": "DGE",
                    "formData": {
                        "BusinessInfo_PersonName_FirstName": "Ike",
                        "BusinessInfo_PersonName_LastName": "Jones",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_AdditionalLocation": 0,
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_CoverageStartDate": "2020-09-22",
                        "BusinessInfo_Locations_Primary_AddrInfo_Addr1": "1419 E SIESTA DR",
                        "BusinessInfo_CommunicationsInfo_EmailInfo_EmailAddr": "test@example.com",
                        "BusinessInfo_CommunicationsInfo_PhoneInfo_PhoneNumber": 6785551212,
                        "BusinessInfo_CommunicationsInfo_PhoneInfo_PhoneExtension": 1100,
                        "BusinessInfo_Locations_Primary_AddrInfo_Country": "USA",
                        "BusinessInfo_Locations_Primary_AddrInfo_StateOrProvCd": "SC",
                        "BusinessInfo_Locations_Primary_AddrInfo_StateOrProvCd": "Florence",
                        "BusinessInfo_CommercialName": "Ijones Company, LLC",
                        "BusinessInfo_Locations_Primary_AddrInfo_PostalCode": 29505,
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_OperatedFromHome": "Yes",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_BusinessOwnershipStructure": "Trust",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_NumOfEmployees": "101",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_SubcontractProfSrvcs": "No",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_AnEConstructionSrvcs": "No",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_SupplyManufactDistbtGoodsOrProducts": "No",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_SimilarInsurance": "No",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_AnEForbiddenProjects1": "No",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_EstmtdPayrollExpense": 500000,
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_EstmtdAnnualGrossSales": 300000,
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_HireNonOwnVehclUse": "No",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_TRIACoverQuoteRq_CoverId": "TRIA",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_AgreeDisagreeToStatements": "I have read and agree with these statements",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_EmailConsent": "Yes",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_RatingInfo_InformationConfirmAgreement": "Yes",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_GeneralLiabilityCoverQuoteRq_CoverId": "CGL",
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_GeneralLiabilityCoverQuoteRq_RatingInfo_AggLOI":   2000000,
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_GeneralLiabilityCoverQuoteRq_RatingInfo_LOI": 1000000,
                        "ProductQuoteRqs_GeneralLiabilityQuoteRq_GeneralLiabilityCoverQuoteRq_RatingInfo_Deductible": 0
                    }
                }
            }
        ]
    }


def convert_to_json():
    for ws in wb:
        ws_data = []
        for index, row in enumerate(ws.iter_rows()):
            if index:
                row_data = gl_data()
                if index % 2 == 0:
                    row_data["via"] = "FIREFOX"
                for column, cell in enumerate(row):
                    insert_cell_value(ws, row_data, cell, column)
                ws_data.append(row_data)
        write_file(ws, ws_data)
    print("done")


def write_file(ws, data):
    filename = ws.title + '.json'
    rel_path = "testdata/bop/" + filename
    abs_file_path = os.path.join(file_dir, rel_path)

    with open(abs_file_path, 'w') as writer:
        writer.write("[\n")
        if data:
            for elem in data:
                writer.write("    ")
                json.dump(elem, writer)
                writer.write(",\n")
            writer.seek(writer.tell() - 2, os.SEEK_SET)
            writer.truncate()
        writer.write("\n]\n")
        writer.truncate()


def store_keys_list():
    ws = wb3["Sheet2"]  # The sheet name
    list_keys = []
    for index, row in enumerate(ws.iter_rows()):
        if not index:
            for cell in row:
                list_keys.append(cell.value)
    return list_keys


def print_header():
    # Print out the header in the first column
    ws = wb3["Sheet2"]  # The sheet name
    for index, row in enumerate(ws.iter_rows()):
        if not index:
            for column, cell in enumerate(row):
                print(str(column) + " - " + cell.value)


if __name__ == '__main__':
    # convert_to_json()
    print_header()
